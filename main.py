"""
Hasta Kayıt Bilgi Sistemi

Geliştirici : Jule Uzun
Teknolojiler : Python, PyQt6, SQLite

Özellikler
- Hasta Ekleme
- Hasta Güncelleme
- Hasta Silme
- Filtreleme
- Excel'e Aktarma
- PDF Raporu
- İstatistik Paneli
"""

# =====================================================
# 1. GEREKLİ KÜTÜPHANELER
# =====================================================
# Standart Kütüphaneler
import os
import sys
import sqlite3
from datetime import datetime

# PyQt6
from PyQt6.QtWidgets import *
from hastakayit import Ui_MainWindow
from istatistik import Ui_Dialog

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# PDF
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Spacer

# Font
pdfmetrics.registerFont(
    TTFont("Arial", "C:/Windows/Fonts/arial.ttf")
    )
pdfmetrics.registerFont(
    TTFont("Arial-Bold", "C:/Windows/Fonts/arialbd.ttf")
    )



# =====================================================
# 2. VERİTABANI
# =====================================================

# Veritabanı dosya yolu
if getattr(sys, "frozen", False):
    klasor = os.path.dirname(sys.executable)
else:
    klasor = os.path.dirname(os.path.abspath(__file__))

db_yolu = os.path.join(klasor, "hastabilgi.db")

# Veritabanı bağlantısını oluştur
baglanti = sqlite3.connect(db_yolu)
imlec = baglanti.cursor()

# Veritabanında hasta tablosunu oluşturur (yoksa oluşturur).
imlec.execute("""
CREATE TABLE IF NOT EXISTS hastabilgi (
    Hasta_TC INTEGER NOT NULL UNIQUE PRIMARY KEY,
    Hasta_Ad TEXT NOT NULL,
    Hasta_Soyad TEXT NOT NULL,
    Telefon TEXT,
    Dogum_Tarihi TEXT,
    Yas INTEGER,
    Poliklinik TEXT NOT NULL,
    Cinsiyet TEXT NOT NULL,
    Kan_Grubu TEXT NOT NULL,
    Durum TEXT NOT NULL
)
""")

imlec.execute("SELECT name FROM sqlite_master WHERE type='table'")

imlec.execute("SELECT COUNT(*) FROM hastabilgi")
sonuc = imlec.fetchone()

baglanti.commit()


# =====================================================
# 3. FONKSİYONLAR
# =====================================================

# 3.1 Yardımcı Fonksiyonlar


# Olumsuz Kullanıcı Girişi Mesajı

def olumsuz_kullanici_girdisi_mesaji():
    mesaj = QMessageBox()
    mesaj.setWindowTitle("Bilgilendirme")
    mesaj.setText("Lütfen Girdiğiniz Değerleri Kontrol Ediniz!")
    mesaj.setStyleSheet("color:darkred;font-size:14px;font-weight:bold")
    mesaj.setIcon(QMessageBox.Icon.Critical)
    mesaj.exec()


# Silme Onayı Mesajı

def silme_onayi_mesaji():
    mesaj = QMessageBox()
    mesaj.setWindowTitle("Onay")
    mesaj.setText("Silmek istediğinize emin misiniz?")
    mesaj.setStyleSheet("color:darkred;font-size:14px;font-weight:bold")
    mesaj.setIcon(QMessageBox.Icon.Question)
    
    mesaj.setStandardButtons(
        QMessageBox.StandardButton.Yes |
        QMessageBox.StandardButton.No
        )
    mesaj.exec()


# Temizleme

def temizle():
    ui.lineEdit_Hasta_TC.clear()
    ui.lineEdit_Hasta_Ad.clear()
    ui.lineEdit_Hasta_Soyad.clear()
    ui.lineEdit_Telefon.clear()
    ui.lineEdit_Dogum_Tarihi.clear()
    ui.lineEdit_Yas.clear()
    
    ui.comboBox_Poliklinik.setCurrentIndex(0)
    ui.comboBox_Cinsiyet.setCurrentIndex(0)
    ui.comboBox_Kan_Grubu.setCurrentIndex(0)

    ui.radioButton_Yatan.setAutoExclusive(False)
    ui.radioButton_Taburcu.setAutoExclusive(False)

    ui.radioButton_Yatan.setChecked(False)
    ui.radioButton_Taburcu.setChecked(False)

    ui.radioButton_Yatan.setAutoExclusive(True)
    ui.radioButton_Taburcu.setAutoExclusive(True)

# Veri Getirme

def veri_getir(index):

    try:

        satir = index.row()

        # Tablodan sadece TC alınır
        tc_item = ui.tableWidget.item(satir, 0)

        if tc_item is None:
            return

        Hasta_TC = tc_item.text()

        # Veritabanından tüm bilgiler alınır
        imlec.execute("""
            SELECT Hasta_TC,
                   Hasta_Ad,
                   Hasta_Soyad,
                   Telefon,
                   Dogum_Tarihi,
                   Yas,
                   Poliklinik,
                   Cinsiyet,
                   Kan_Grubu,
                   Durum
            FROM hastabilgi
            WHERE Hasta_TC = ?
        """, (Hasta_TC,))

        kayit = imlec.fetchone()

        if not kayit:
            return

        ui.lineEdit_Hasta_TC.setText(str(kayit[0]))
        ui.lineEdit_Hasta_Ad.setText(kayit[1])
        ui.lineEdit_Hasta_Soyad.setText(kayit[2])
        ui.lineEdit_Telefon.setText(kayit[3] or "")
        ui.lineEdit_Dogum_Tarihi.setText(kayit[4] or "")
        ui.lineEdit_Yas.setText(str(kayit[5]) or "")
        ui.comboBox_Poliklinik.setCurrentText(kayit[6])
        ui.comboBox_Cinsiyet.setCurrentText(kayit[7])
        ui.comboBox_Kan_Grubu.setCurrentText(kayit[8])

        if kayit[9] == "Yatan":
            ui.radioButton_Yatan.setChecked(True)
        else:
            ui.radioButton_Taburcu.setChecked(True)

    except Exception as hata:
        ui.statusbar.showMessage(f"Hata: {hata}", 3000)
        
        
# Yaş Hesaplama

def yas_hesapla(dogum_tarihi):
    try:
        dogum = datetime.strptime(dogum_tarihi, "%d.%m.%Y")
        bugun = datetime.today()

        yas = bugun.year - dogum.year

        if (bugun.month, bugun.day) < (dogum.month, dogum.day):
            yas -= 1

        return yas

    except ValueError:
        return None


# 3.2 Veri Ekleme

#Hastaya ait bilgileri doğrular ve veritabanına yeni kayıt ekler.

def ekle():

    try:

        tc = ui.lineEdit_Hasta_TC.text().strip()

        if not tc:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Hasta TC boş bırakılamaz."
            )
            return

        Hasta_TC = int(tc)
        
        Hasta_Ad = ui.lineEdit_Hasta_Ad.text().strip()
        Hasta_Soyad = ui.lineEdit_Hasta_Soyad.text().strip()
        
        if not Hasta_Ad or not Hasta_Soyad:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Ad ve Soyad boş bırakılamaz."
            )
            return
        
        Telefon = ui.lineEdit_Telefon.text().strip()
        Dogum_Tarihi = ui.lineEdit_Dogum_Tarihi.text().strip()

        Yas = yas_hesapla(Dogum_Tarihi)

        if Yas is None:
            QMessageBox.warning(pencere,"Uyarı","Doğum tarihini GG.AA.YYYY formatında giriniz!")
            return

        ui.lineEdit_Yas.setText(str(Yas))

        Poliklinik = ui.comboBox_Poliklinik.currentText()
        Cinsiyet = ui.comboBox_Cinsiyet.currentText()
        Kan_Grubu = ui.comboBox_Kan_Grubu.currentText()
        
        if Poliklinik == "Seçiniz...":
            QMessageBox.warning(pencere, "Uyarı", "Lütfen poliklinik seçiniz!")
            return

        if Cinsiyet == "Seçiniz...":
            QMessageBox.warning(pencere, "Uyarı", "Lütfen cinsiyet seçiniz!")
            return

        if Kan_Grubu == "Seçiniz...":
            QMessageBox.warning(pencere, "Uyarı", "Lütfen kan grubu seçiniz!")
            return

        if ui.radioButton_Yatan.isChecked():
            Durum = "Yatan"
        elif ui.radioButton_Taburcu.isChecked():
            Durum = "Taburcu"
        else:
            QMessageBox.warning(pencere, "Uyarı", "Lütfen durum seçiniz!")
            return

        
        sorgu = """
        INSERT INTO hastabilgi (
            Hasta_TC,
            Hasta_Ad,
            Hasta_Soyad,
            Telefon,
            Dogum_Tarihi,
            Yas,
            Poliklinik,
            Cinsiyet,
            Kan_Grubu,
            Durum
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        imlec.execute(
            sorgu,
            (
                Hasta_TC,
                Hasta_Ad,
                Hasta_Soyad,
                Telefon,
                Dogum_Tarihi,
                Yas,
                Poliklinik,
                Cinsiyet,
                Kan_Grubu,
                Durum
            )
        )

        baglanti.commit()

        tumunu_listele()
        temizle()

        ui.lineEdit_Hasta_TC.setFocus()

        ui.statusbar.showMessage(
            "Hasta kaydı başarıyla eklendi.",
            3000   
        )

    except sqlite3.IntegrityError:
        QMessageBox.warning(
        pencere,
        "Uyarı",
        "Bu TC numarası zaten kayıtlı."
    )

    except ValueError:
        QMessageBox.warning(
        pencere,
        "Uyarı",
        "TC yalnızca rakamlardan oluşmalıdır."
    )

    except Exception as hata:
        QMessageBox.critical(
        pencere,
        "Beklenmeyen Hata",
        str(hata)
    )
        

# 3.3 Veri Silme

# Seçilen hasta kaydını siler.
def veri_silme():
    silme_onayı = QMessageBox.question(
        pencere,
        "Silme Onayı",
        "Silmek İstediğinize Emin misiniz?",
        QMessageBox.StandardButton.Yes |
        QMessageBox.StandardButton.No
    )

    if silme_onayı != QMessageBox.StandardButton.Yes:
        return
    
    row = ui.tableWidget.currentRow()

    if row < 0:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Lütfen satır seçin!"
            )
            return

    tc_item = ui.tableWidget.item(row, 0)

    if tc_item is None:
        QMessageBox.warning(
            pencere,
            "Uyarı",
            "Geçersiz kayıt."
        )
        return

    Hasta_TC = tc_item.text()

    try:
        imlec.execute(
            "DELETE FROM hastabilgi WHERE Hasta_TC = ?",
            (Hasta_TC,)
        )
        baglanti.commit()
        tumunu_listele()
        temizle()
            
        ui.statusbar.showMessage(
        "Hasta kaydı başarıyla silindi.",
        3000
        )    

        
    except Exception as hata:
            
        QMessageBox.critical(
            pencere,
            "Silme Hatası",
            str(hata)
        )
        ui.statusbar.showMessage(str(hata),3000)

# 3.4 Veri Güncelleme

# Seçilen kaydını günceller.
def veri_guncelle():
    
    try:
        Hasta_TC = ui.lineEdit_Hasta_TC.text().strip()

        if not Hasta_TC:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Lütfen güncellenecek hastayı seçiniz."
            )
            return
        
        Hasta_TC = int(Hasta_TC)
        
        # Hasta Bilgileri
        Hasta_Ad = ui.lineEdit_Hasta_Ad.text().strip()
        Hasta_Soyad = ui.lineEdit_Hasta_Soyad.text().strip()
        Telefon = ui.lineEdit_Telefon.text().strip()
        Dogum_Tarihi = ui.lineEdit_Dogum_Tarihi.text().strip()
        
        
        # Ad Soyad Kontrolü
        if not Hasta_Ad or not Hasta_Soyad:
            QMessageBox.warning(
                 pencere,
                "Uyarı",
                "Ad ve Soyad boş bırakılamaz."
            )
            return
        
        # Yaş Hesaplama
        Yas = yas_hesapla(Dogum_Tarihi)
    
        if Yas is None:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Doğum tarihini GG.AA.YYYY formatında giriniz!"
            )
            return
    
        # ComboBox Bilgileri
        Poliklinik = ui.comboBox_Poliklinik.currentText()
        Cinsiyet = ui.comboBox_Cinsiyet.currentText()
        Kan_Grubu = ui.comboBox_Kan_Grubu.currentText()
        
        if Poliklinik == "Seçiniz...":
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Lütfen poliklinik seçiniz!"
            )
            return

        if Cinsiyet == "Seçiniz...":
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Lütfen cinsiyet seçiniz!"
            )
            return

        if Kan_Grubu == "Seçiniz...":
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Lütfen kan grubu seçiniz!"
            )
            return
        
        # Hasta Durumu
        if ui.radioButton_Yatan.isChecked():
            Durum = "Yatan"

        elif ui.radioButton_Taburcu.isChecked():
            Durum = "Taburcu"

        else:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Lütfen durum seçiniz!"
            )
            return

        # Güncelleme Sorgusu
        sorgu = """
        UPDATE hastabilgi
        SET
            Hasta_Ad=?,
            Hasta_Soyad=?,
            Telefon=?,
            Dogum_Tarihi=?,
            Yas=?,
            Poliklinik=?,
            Cinsiyet=?,
            Kan_Grubu=?,
            Durum=?
        WHERE Hasta_TC=?
        """
    
        imlec.execute(
            sorgu,
            (
                Hasta_Ad,
                Hasta_Soyad,
                Telefon,
                Dogum_Tarihi,
                Yas,
                Poliklinik,
                Cinsiyet,
                Kan_Grubu,
                Durum,
                Hasta_TC
            )
        )

        baglanti.commit()
    
        tumunu_listele()
        temizle()
    
        ui.statusbar.showMessage(
            "Hasta kaydı başarıyla güncellendi.",
            3000
        )
    
    except ValueError:
        QMessageBox.warning(
            pencere,
            "Uyarı",
            "Hasta TC yalnızca rakamlardan oluşmalıdır."
        )

    except Exception as hata:
        QMessageBox.critical(
            pencere,
            "Güncelleme Hatası",
            str(hata)
        )

    
  
# 3.5 Verileri Listeleme

# Veritabanındaki tüm kayıtları tabloya listeler.
def tumunu_listele():
    try:
        
        sorgu = """
        SELECT
            Hasta_TC,
            Hasta_Ad,
            Hasta_Soyad,
            Telefon,
            Yas,
            Poliklinik,
            Kan_Grubu,
            Durum
        FROM hastabilgi
        """
        imlec.execute(sorgu)
        
        veriler = imlec.fetchall()
        
        ui.tableWidget.clearContents()
        ui.tableWidget.setRowCount(len(veriler))

        for satir, kayit in enumerate(veriler):
            
            for sutun, veri in enumerate(kayit):
                
                ui.tableWidget.setItem(
                    satir,
                    sutun,
                    QTableWidgetItem(str(veri))
                )

    except Exception as hata:
        QMessageBox.critical(
            pencere,
            "Listeleme Hatası",
            str(hata)
        )

        
# 3.6 Filtreleme

# TC, ad, soyad, poliklinik ve kan grubuna göre arama yapar.

def filtrele():

    tc = ui.lineEdit_Arama_TC.text().strip()
    ad = ui.lineEdit_Arama_Hasta_Ad.text().strip()
    soyad = ui.lineEdit_Arama_Hasta_Soyad.text().strip()
    poliklinik = ui.comboBox_Poliklinik_Filtrele.currentText()
    kan = ui.comboBox_Kan_Grubu_Filtrele.currentText()

    sorgu = """
    SELECT 
        Hasta_TC,
        Hasta_Ad,
        Hasta_Soyad,
        Telefon,
        Yas,
        Poliklinik,
        Kan_Grubu,
        Durum
    FROM hastabilgi
    WHERE 1=1
    """

    parametreler = []

    if tc != "":
        sorgu += " AND CAST(Hasta_TC AS TEXT) LIKE ?"
        parametreler.append(f"%{tc}%")

    if ad != "":
        sorgu += " AND Hasta_Ad LIKE ?"
        parametreler.append(f"%{ad}%")

    if soyad != "":
        sorgu += " AND Hasta_Soyad LIKE ?"
        parametreler.append(f"%{soyad}%")

    if poliklinik != "Seçiniz...":
        sorgu += " AND Poliklinik=?"
        parametreler.append(poliklinik)

    if kan != "Seçiniz...":
        sorgu += " AND Kan_Grubu=?"
        parametreler.append(kan)

    try:

        imlec.execute(sorgu, parametreler)

        veriler = imlec.fetchall()
        
        ui.tableWidget.clearContents()
        ui.tableWidget.setRowCount(0)
        ui.tableWidget.setRowCount(len(veriler))

        for satir, kayit in enumerate(veriler):
            for sutun, veri in enumerate(kayit):
                ui.tableWidget.setItem(
                    satir,
                    sutun,
                    QTableWidgetItem(str(veri))
                )
                
        ui.statusbar.showMessage(
            f"{len(veriler)} kayıt bulundu.",
            3000
        )
        
    except Exception as hata:
        QMessageBox.critical(
            pencere,
            "Filtreleme Hatası",
            str(hata)
        )
        
        
# 3.7 Excel'e Aktarma

def excel_aktar():
    try:
        dosya_adi, _ = QFileDialog.getSaveFileName(
            pencere,
            "Excel Dosyasını Kaydet",
            "Hasta_Kayit_Listesi.xlsx",
            "Excel Dosyası (*.xlsx)"
        )

        if not dosya_adi:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Hasta Listesi"

    # Başlıkları yazar
        
        for sutun in range(ui.tableWidget.columnCount()):
            baslik = ui.tableWidget.horizontalHeaderItem(sutun)

            if baslik:
                hucre = ws.cell(row=1, column=sutun + 1)
                hucre.value = baslik.text()
                hucre.font = Font(bold=True)


    # Verileri yazar
        
        for satir in range(ui.tableWidget.rowCount()):
            for sutun in range(ui.tableWidget.columnCount()):
                item = ui.tableWidget.item(satir, sutun)

                if item:
                    ws.cell(
                        row=satir + 2,
                        column=sutun + 1
                    ).value = item.text()
    

    # Sütun genişliklerini otomatik ayarlar
        
        for column_cells in ws.columns:
            uzunluk = 0
            harf = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    uzunluk = max(uzunluk, len(str(cell.value)))

            ws.column_dimensions[harf].width = uzunluk + 3

            wb.save(dosya_adi)

        QMessageBox.information(
            pencere,
            "Başarılı",
            "Excel dosyası başarıyla oluşturuldu."
        )    

    except Exception as hata:
        QMessageBox.critical(
            pencere,
            "Excel Hatası",
            str(hata)
        )
        
        
def sayfa_numarasi(canvas, doc):
    canvas.saveState()
    canvas.setFont("Arial", 9)

    canvas.drawRightString(
        19 * cm,
        1 * cm,
        f"Sayfa {doc.page}"
    )

    canvas.restoreState()
        
# 3.8 PDF'e Aktarma

def pdf_aktar():
    try:
        varsayilan = f"Hasta_Kayit_Raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        dosya_adi, _ = QFileDialog.getSaveFileName(
            pencere,
            "PDF Kaydet",
            varsayilan,
            "PDF Dosyası (*.pdf)"
        )
        
        if not dosya_adi:
            return
        
        if ui.tableWidget.rowCount() == 0:
            QMessageBox.warning(
                pencere,
                "Uyarı",
                "Aktarılacak kayıt bulunamadı."
            )
            return
        
        from reportlab.lib.pagesizes import A4

        belge = SimpleDocTemplate(
            dosya_adi,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )

        stiller = getSampleStyleSheet()

        baslik = Paragraph(
        "<font face='Arial-Bold' size='16'>HASTA KAYIT RAPORU</font>",
        stiller["Title"]
        )

        olusturma_tarihi = datetime.now().strftime("%d.%m.%Y %H:%M")

        tarih = Paragraph(
        f"<font face='Arial'>Oluşturulma Tarihi : {olusturma_tarihi}</font>",
        stiller["Normal"]
        )

    # Toplam kayıt sayısı
        toplam = Paragraph(
            f"<font face='Arial'>Toplam Hasta Sayısı : {ui.tableWidget.rowCount()}</font>",
            stiller["Normal"]
        )       
        veri = []


    # Başlıklar
    
        basliklar = []

        for i in range(ui.tableWidget.columnCount()):
            basliklar.append(ui.tableWidget.horizontalHeaderItem(i).text())

        veri.append(basliklar)


    # Tablodaki veriler
    
        for satir in range(ui.tableWidget.rowCount()):

            satir_verisi = []

            for sutun in range(ui.tableWidget.columnCount()):

                item = ui.tableWidget.item(satir, sutun)

                satir_verisi.append(item.text() if item else "")

            veri.append(satir_verisi)

        tablo = Table(
            veri,
            repeatRows=1
        )

        tablo.setStyle(TableStyle([

            ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),

            ('FONTNAME', (0,0), (-1,-1), 'Arial'),

            ('GRID', (0,0), (-1,-1), 1, colors.black),

            ('BACKGROUND', (0,1), (-1,-1), colors.beige),

            ('ALIGN', (0,0), (-1,-1), 'CENTER'),

            ('BOTTOMPADDING', (0,0), (-1,0), 10),
            ('FONTSIZE',(0,0),(-1,-1),9),

        ]))

        icerik = [
            baslik,
            Spacer(1, 0.4 * cm),
            tarih,
            Spacer(1, 0.2 * cm),
            toplam,
            Spacer(1, 0.5 * cm),
            tablo
        ]

        belge.build(
            icerik,
            onFirstPage=sayfa_numarasi,
            onLaterPages=sayfa_numarasi
        )

   

        QMessageBox.information(
            pencere,
            "Başarılı",
            "PDF başarıyla oluşturuldu."
        )
    
    except Exception as hata:
        QMessageBox.critical(
            pencere,
            "PDF Hatası",
            str(hata)
        )
    
# 3.9 İstatistik Penceresi

# Hasta istatistiklerini ayrı bir pencerede gösterir.
class IstatistikPenceresi(QDialog):

    def __init__(self):
        super().__init__()

        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.istatistikleri_yukle()

        self.ui.btn_Kapat.clicked.connect(self.close)


    def istatistikleri_yukle(self):
        try:
            
            # Toplam Hasta
            imlec.execute("SELECT COUNT(*) FROM hastabilgi")
            toplam = imlec.fetchone()[0]

            # Yatan
            imlec.execute("SELECT COUNT(*) FROM hastabilgi WHERE Durum='Yatan'")
            yatan = imlec.fetchone()[0]

            # Taburcu
            imlec.execute("SELECT COUNT(*) FROM hastabilgi WHERE Durum='Taburcu'")
            taburcu = imlec.fetchone()[0]

            # Erkek
            imlec.execute("SELECT COUNT(*) FROM hastabilgi WHERE Cinsiyet='Erkek'")
            erkek = imlec.fetchone()[0]

            # Kadın
            imlec.execute("SELECT COUNT(*) FROM hastabilgi WHERE Cinsiyet='Kadın'")
            kadin = imlec.fetchone()[0]

            # Ortalama Yaş
            imlec.execute("SELECT AVG(Yas) FROM hastabilgi")
            ort_yas = imlec.fetchone()[0]

            if ort_yas is None:
                ort_yas = 0
            else:
                ort_yas = round(ort_yas, 1)

            # Poliklinik Sayısı
            imlec.execute("SELECT COUNT(DISTINCT Poliklinik) FROM hastabilgi")
            poliklinik = imlec.fetchone()[0]

            # Kan Grupları
            imlec.execute("""
                SELECT Kan_Grubu, COUNT(*)
                FROM hastabilgi
                GROUP BY Kan_Grubu
                ORDER BY Kan_Grubu
            """)

            sonuc = imlec.fetchall()

            kanlar = ""

            for grup, adet in sonuc:
                kanlar += f"{grup} : {adet}\n"
                
            self.ui.textEdit_KanGruplari.setPlainText(kanlar)
            

            # Label'lara Yazar

            self.ui.label_Toplam.setText(f"👥 Toplam Hasta : {toplam}")

            self.ui.label_Yatan.setText(f"🟢 Yatan : {yatan}")

            self.ui.label_Taburcu.setText(f"🔵 Taburcu : {taburcu}")

            self.ui.label_Erkek.setText(f"👨 Erkek : {erkek}")

            self.ui.label_Kadin.setText(f"👩 Kadın : {kadin}")

            self.ui.label_OrtalamaYas.setText(f"🎂 Ortalama Yaş : {ort_yas}")

            self.ui.label_Poliklinik.setText(f"🏥 Poliklinik Sayısı : {poliklinik}")

        except Exception as hata:

            QMessageBox.critical(
                self,
                "İstatistik Hatası",
                str(hata)
            )


    # Global Değişkenler
           
    istatistik_penceresi = None

def istatistik_goster():

    global istatistik_penceresi

    istatistik_penceresi = IstatistikPenceresi()

    istatistik_penceresi.exec()
    
    
# =====================================================
# 4. UYGULAMANIN OLUŞTURULMASI
# =====================================================    

# QApplication oluştur
uygulama = QApplication(sys.argv)

# Ana pencere oluştur
pencere = QMainWindow()

# Arayüzü yükle
ui = Ui_MainWindow()
ui.setupUi(pencere)

# Pencere başlığı
pencere.setWindowTitle("Hasta Kayıt Sistemi v1.0")

# =====================================================
# 5. ARAYÜZ AYARLARI
# =====================================================

# TableWidget ayarları

ui.tableWidget.setColumnCount(8)
ui.tableWidget.setHorizontalHeaderLabels([
    "Hasta TC",
    "Ad",
    "Soyad",
    "Telefon",
    "Yaş",
    "Poliklinik",
    "Kan Grubu",
    "Durum"
])


ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

ui.tableWidget.setStyleSheet("""
QHeaderView::section{
    background-color:lightblue;
    color:black;
    font-weight:bold;
    border:1px solid #B5C8D6;
}

QTableWidget::item:selected{
    background-color:#7DB9DE;
    color:black;
}
""")

ui.tableWidget.setSelectionBehavior(
    QTableWidget.SelectionBehavior.SelectRows
)

ui.tableWidget.setSelectionMode(
    QTableWidget.SelectionMode.SingleSelection
)

ui.tableWidget.setEditTriggers(
    QAbstractItemView.EditTrigger.NoEditTriggers
)


# StatusBar

ui.statusbar.showMessage("Veritabanı bağlantısı başarıyla kuruldu.", 3000)
ui.statusbar.setStyleSheet("color:green;font-size:18px;font-weight:bold;")


# =====================================================
# 6. SİNYAL (BUTON) BAĞLANTILARI
# =====================================================

ui.btn_Ekle.clicked.connect(ekle)
ui.btn_Sil.clicked.connect(veri_silme)
ui.btn_Guncelle.clicked.connect(veri_guncelle)
ui.tableWidget.clicked.connect(veri_getir)
ui.btn_TumunuListele.clicked.connect(tumunu_listele)
ui.btn_Filtrele.clicked.connect(filtrele)
ui.pushButton_Excel.clicked.connect(excel_aktar)
ui.pushButton_PDF.clicked.connect(pdf_aktar)
ui.pushButton_Istatistik.clicked.connect(istatistik_goster)


# =====================================================
# 7. UYGULAMANIN BAŞLATILMASI
# =====================================================

pencere.show()

sys.exit(uygulama.exec())


